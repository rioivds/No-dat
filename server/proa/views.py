import datetime
from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from proa.models import Curso
from django.shortcuts import get_object_or_404
from django.shortcuts import render
from django.http import HttpResponseRedirect
from openpyxl import Workbook
from .forms import ImportarAlumnosForm
from .importar_alumnos import importar_alumnos  # Importa la función que creamos antes
import openpyxl
from django.http import HttpResponse
from django.db.models import F
from .models import Alumno, Curso, Materia, Profesor, Calificaciones

TEMPLATE_DIR = ('os.path.join(BASE_DIR,"templates")')

def importar_materias(request):
    mensaje = ''  # Inicializar la variable mensaje

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo_excel = request.FILES['archivo_excel']
        try:
            workbook = openpyxl.load_workbook(archivo_excel)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nombre = row[0]
                horas_catedra = row[1]
                horario = row[2]
                profesor_nombre = row[3]  # Suponiendo que en la columna 4 del Excel tienes el nombre del profesor
                curso_anio = row[4]  # Suponiendo que en la columna 5 del Excel tienes el año del curso

                # Obtener o crear el profesor
                profesor, _ = Profesor.objects.get_or_create(nombre=profesor_nombre)

                # Obtener el curso
                curso, _ = Curso.objects.get_or_create(anio=curso_anio)

                # Crear la materia
                materia = Materia(nombre=nombre, horas_catedra=horas_catedra, horario=horario, profesor=profesor, curso=curso)
                materia.save()

            mensaje = 'Materias importadas correctamente.'
        except Exception as e:
            mensaje = f'Error al importar las materias: {str(e)}'

    return render(request, 'materias/importar_materias.html', {'mensaje': mensaje})


def importar_alumnos_view(request):
    if request.method == 'POST':
        form = ImportarAlumnosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            importar_alumnos(archivo_excel)
            return HttpResponseRedirect('/alumnos')  # Redirigir a la página principal después de la importación
    else:
        form = ImportarAlumnosForm()
    return render(request, 'alumnos/importar_alumnos.html', {'form': form})

def exportar_alumnos(request):
    alumnos = Alumno.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Alumnos'

    # Encabezados de las columnas
    sheet['A1'] = 'DNI'
    sheet['B1'] = 'Nombre'
    sheet['C1'] = 'Apellido'
    sheet['D1'] = 'Fecha de Nacimiento'
    sheet['E1'] = 'Email'
    sheet['F1'] = 'Repitio'
    sheet['G1'] = 'Año'

    # Llenar el contenido con los datos de los alumnos
    for index, alumno in enumerate(alumnos, start=2):
        sheet[f'A{index}'] = alumno.dni
        sheet[f'B{index}'] = alumno.nombre
        sheet[f'C{index}'] = alumno.apellido
        sheet[f'D{index}'] = alumno.fecha_nacimiento.strftime('%Y-%m-%d')
        sheet[f'E{index}'] = alumno.email
        sheet[f'F{index}'] = 'Si' if alumno.repitio else 'No'
        sheet[f'G{index}'] = alumno.curso.anio

    # Agregar estilo a los encabezados
    header_font = openpyxl.styles.Font(bold=True)
    for cell in sheet['1']:
        cell.font = header_font

    # Generar la respuesta con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=alumnos.xlsx'
    workbook.save(response)
    return response

def importar_calificaciones(request):
    mensaje = ''  # Asignar un valor predeterminado o una cadena vacía

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo_excel = request.FILES['archivo_excel']
        try:
            workbook = openpyxl.load_workbook(archivo_excel)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                alumno_dni = row[0]  # Suponiendo que en la columna 1 del Excel tienes el DNI del alumno
                curso_id = row[1]  # Suponiendo que en la columna 2 del Excel tienes el ID del curso
                materia_id = row[2]  # Suponiendo que en la columna 3 del Excel tienes el ID de la materia
                profesor_id = row[3]  # Suponiendo que en la columna 4 del Excel tienes el ID del profesor
                fecha = row[4]  # Suponiendo que en la columna 5 del Excel tienes la fecha de la calificación
                nota = row[5]  # Suponiendo que en la columna 6 del Excel tienes la nota de la calificación
                final = row[6]  # Suponiendo que en la columna 7 del Excel tienes un valor booleano para indicar si es final

                # Obtener o crear el alumno
                alumno, _ = Alumno.objects.get_or_create(dni=alumno_dni)

                # Obtener o crear la materia
                materia, _ = Materia.objects.get_or_create(id=materia_id)

                # Obtener o crear el profesor
                profesor, _ = Profesor.objects.get_or_create(dni=profesor_id)

                # Obtener o crear el curso
                curso, _ = Curso.objects.get_or_create(id=curso_id)

                # Crear la calificación
                calificacion = Calificaciones(alumno=alumno, curso=curso, materia=materia, profesor=profesor, fecha=fecha, nota=nota, final=final)
                calificacion.save()

            mensaje = 'Calificaciones importadas correctamente.'
        except Exception as e:
            mensaje = f'Error al importar las calificaciones: {str(e)}'

    return render(request, 'calificaciones/importar_calificaciones.html', {'mensaje': mensaje})

def index(request):
    cursos = Curso.objects.all()
    today = datetime.datetime.now()
    return render(request, 'index.html',
                  {"today": today,
                   "titulo": ' Titulo desde la vista',
                   'texto': 'aca ponemos algo de texto desde la base',
                   'cursos': cursos
                   })

def guardar(request):
    return HttpResponse('Hola Sou guardar')